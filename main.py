import util
import openpyxl
from pathlib import Path
import csv

from clinicalTrials import get_trial_info, retreive_trials
from references import write_references


def read_grants_file():
    xlsx_file = Path('data', 'SearchResult_Export_27Jan2021_103910.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)
    grants = []
    for row in sheet.iter_rows(min_row=2):
        grant = {}
        for cell in row:
            grant[col_names[cell.col_idx-1]] =  util.replace_characters(cell.value)
        grants.append(grant)
    return grants


# createFundedOrganization(
# city: String!
# country: String!
# id: ID!
# name: String!
# organizationID: String!
# state: String!
# type: String
#Organization ID (IPF)	Organization Name	Organization City	Organization State	Organization Type	Organization Zip	Organization Country

def create_FundedOrganization(grant,server):
    graph_id: str = util.get_unique_id('organization_')
    lc = grant["Organization Name"] + ' ' + grant["Organization City"] + ' ' + grant["Organization State"]
    lc = lc.lower()
    s: str = f'{graph_id}: createFundedOrganization( city: \\"{grant["Organization City"]}\\",country: \\"{grant["Organization Country"]}\\", id: \\"{graph_id}\\",' \
             f'name:\\"{grant["Organization Name"]}\\",organizationID:\\"{grant["Organization ID (IPF)"]}\\",state:\\"{grant["Organization State"]}\\",type:\\"{grant["Organization Type"]}\\", lower_case_search_string: \\"{lc}\\" ),'
    # print(graph_id)
    util.send_mutation(s, server)
    # print(graph_id, flush=True)

    return graph_id

# createNIHInstituteOrCenter(id: ID!name: String!)
def create_NIHInstituteOrCenter(name,server):
    graph_id: str = util.get_unique_id('NIH_IC_')

    s: str = f'{graph_id}: createNIHInstituteOrCenter( id: \\"{graph_id}\\", name:\\"{name}\\", lower_case_search_string: \\"{name.lower()}\\" ),'
    # print(graph_id)
    util.send_mutation(s, server)
    # print(graph_id, flush=True)
    return graph_id


def create_NIHGrant(grant,server,oranization_dict,ic_dict,pi_dict,core_project_dict):
    graph_id: str = util.get_unique_id('grant_')
    if grant['Total Cost'] == '':
        grant['Total Cost'] = 0
    if grant['Total Cost IC'] == '':
        grant['Total Cost IC'] = 0
    if grant['Direct Cost IC'] == '':
        grant['Direct Cost IC'] = 0
    if grant['InDirect Cost IC'] == '':
        grant['InDirect Cost IC'] = 0
    if grant['Total Cost (Sub Projects)'] == '':
        grant['Total Cost (Sub Projects)'] = 0
    # lower_case_search_string = grant["Project Terms"] + ' ' + grant['Project Title']+ ' ' + grant["Contact PI / Project Leader"]
    lower_case_search_string =   grant['Project Title']+ ' ' + grant["Contact PI / Project Leader"]
    lower_case_search_string = lower_case_search_string.lower()
    # if len(grant["Project Abstract"]) > 3300:
    #     print(len(grant["Project Abstract"]), flush=True)

    # grant["Project Abstract"] = ''
    grant["Project Terms"] = ''

    s: str = f'{graph_id}: createNIHGrant(CFDACode: \\"{grant["CFDA Code"]}\\",  FOA: \\"{grant["FOA"]}\\",  IC: \\"{grant["IC"]}\\", NIHCOVID19Response:  \\"{grant["NIH COVID-19 Response"]}\\", abstract:\\"{grant["Project Abstract"]}\\",' \
             f'activity:\\"{grant["Activity"]}\\",  applicationID:\\"{grant["Application ID"]}\\", awardNoticeDate:\\"{fix_dates(grant["Award Notice Date"])}\\", budgetEndDate:\\"{fix_dates(grant["Budget End Date"])}\\", budgetStartDate:\\"{fix_dates(grant["Budget Start Date"])}\\", ' \
             f'department:\\"{grant["Department"]}\\", directCost_IC: {grant["Direct Cost IC"]},fiscalYear:\\"{grant["Fiscal Year"]}\\",  fundingMechanism:\\"{grant["Funding Mechanism"]}\\",  ' \
             f'id: \\"{graph_id}\\", inDirectCost_IC: {grant["InDirect Cost IC"]}, nihSpendingCategorization: \\"{grant["NIH Spending Categorization"]}\\",  otherPIorProjectLeaders: \\"{grant["Other PI or Project Leader(s)"]}\\", ' \
             f'programOfficialInformation: \\"{grant["Program Official Information"]}\\", projectEndDate: \\"{fix_dates(grant["Project End Date"])}\\", projectNumber: \\"{grant["Project Number"]}\\", projectStartDate: \\"{fix_dates(grant["Project Start Date"])}\\", projectTerms: \\"{grant["Project Terms"]}\\", ' \
             f'projectTitle: \\"{grant["Project Title"]}\\", publicHealthRelevance: \\"{grant["Public Health Relevance"]}\\", serialNumber: \\"{grant["Serial Number"]}\\", studySection: \\"{grant["Study Section"]}\\", supportYear: \\"{grant["Support Year"]}\\", ' \
             f'totalCost: {grant["Total Cost"]}, totalCost_IC: {grant["Total Cost IC"]}, totalCost_SubProjects: {grant["Total Cost (Sub Projects)"]}, type:\\"{grant["Type"]}\\", lower_case_search_string:\\"{lower_case_search_string}\\" ),'
    # print(graph_id)
    # if grant["Project Number"]=='5P30AG010161-30':
    #     print('here',flush=True)
    util.send_mutation(s, server)
    master_lower_case_search_string = lower_case_search_string

    master_lower_case_search_string = handle_organization(grant, graph_id, oranization_dict, master_lower_case_search_string,server)
    master_lower_case_search_string = handle_ICs(grant, graph_id, ic_dict, master_lower_case_search_string,server)
    master_lower_case_search_string = handle_pi(grant,graph_id,pi_dict,master_lower_case_search_string,server)
    handle_core_project(grant,graph_id,core_project_dict,master_lower_case_search_string,server)

    return graph_id


# createPrincipalInvestigator(
# firstName: String!
# id: ID!
# middleName: String
# personID: String!
# surname: String!)
def handle_pi(grant,graph_id,pi_dict,master_lower_case_search_string,server):
    contactPI = grant["Contact PI / Project Leader"]
    master_lower_case_search_string += contactPI
    if not contactPI in pi_dict:
        contactPI_graph_id: str = util.get_unique_id('PI_')
        toks = contactPI.split(',')
        last_name = toks[0]
        toks1 = toks[1].strip().split(' ')
        first_name = toks1[0]
        if len(toks1)>1:
            middle_name = toks1[1]
        else:
            middle_name = ''

        person_id = grant["Contact PI Person ID"]
        lc = first_name.lower() + ' ' + last_name.lower()
        if (middle_name == ''):
            s = f'{contactPI_graph_id}: createPrincipalInvestigator( firstName:\\"{first_name}\\", id: \\"{contactPI_graph_id}\\", personID: \\"{person_id}\\", surname:\\"{last_name}\\", lower_case_search_string: \\"{lc}\\"),'
        else:
            s = f'{contactPI_graph_id}: createPrincipalInvestigator( firstName:\\"{first_name}\\", id: \\"{contactPI_graph_id}\\", middleName: \\"{middle_name}\\",personID: \\"{person_id}\\", surname:\\"{last_name}\\", lower_case_search_string: \\"{lc}\\" ),'
        # print(contactPI_graph_id)
        util.send_mutation(s, server)
        pi_dict[contactPI] = contactPI_graph_id
#         addNIHGrantContactPIorProjectLeader(contactPIorProjectLeader: [ID!]!id: ID!)
    contactPI_graph_id = pi_dict[contactPI]
    s = f'addNIHGrantContactPIorProjectLeader(id: \\"{graph_id}\\", contactPIorProjectLeader: [\\"{contactPI_graph_id}\\"])'
    util.send_mutation(s, server)
    return master_lower_case_search_string


def handle_ICs(grant, graph_id, ic_dict, master_lower_case_search_string,server):
    administeringIC = grant["Administering IC"]
    master_lower_case_search_string += administeringIC
    if not administeringIC in ic_dict:
        administeringIC_graph_id = create_NIHInstituteOrCenter(administeringIC, server)
        ic_dict[administeringIC] = administeringIC_graph_id
    administeringIC_graph_id = ic_dict[administeringIC]
    s = f'addNIHGrantAdministeringIC(id: \\"{graph_id}\\", administeringIC: [\\"{administeringIC_graph_id}\\"])'
    util.send_mutation(s, server)
    fundingIC = grant["Funding IC(s)"]
    master_lower_case_search_string += fundingIC
    if not fundingIC in ic_dict:
        fundingIC_graph_id = create_NIHInstituteOrCenter(fundingIC, server)
        ic_dict[fundingIC] = fundingIC_graph_id
    fundingIC_graph_id = ic_dict[fundingIC]
    s = f'addNIHGrantFundingIC(id: \\"{graph_id}\\", fundingIC: [\\"{fundingIC_graph_id}\\"])'
    util.send_mutation(s, server)
    return master_lower_case_search_string


def handle_organization(grant, graph_id, oranization_dict, master_lower_case_search_string, server):
    organization_name:str = grant["Organization Name"]
    master_lower_case_search_string += organization_name
    if not organization_name in oranization_dict:
        organization_graph_id = create_FundedOrganization(grant, server)
        oranization_dict[organization_name] = organization_graph_id
    organization_graph_id = oranization_dict[organization_name]
    s = f'addNIHGrantOrganization(id: \\"{graph_id}\\", organization: [\\"{organization_graph_id}\\"])'
    util.send_mutation(s, server)
    return master_lower_case_search_string
    # print(organization_graph_id,flush=True)


# createCoreProject(coreProjectNumber: String!id: ID!)
def handle_core_project(grant, graph_id, core_project_dict, master_lower_case_search_string, server):
    if (grant["Serial Number"]!=''):
        core_project_number = grant["Activity"] + grant["IC"] +	grant["Serial Number"]
        master_lower_case_search_string += core_project_number
        if core_project_number not in core_project_dict:
            core_project_graph_id: str = util.get_unique_id('core_project_')
            core_project_dict[core_project_number] = core_project_graph_id
            s = f'createCoreProject(id: \\"{core_project_graph_id}\\", coreProjectNumber: \\"{core_project_number}\\", lower_case_search_string: \\"{master_lower_case_search_string.lower()}\\")'
            util.send_mutation(s, server)
        core_project_graph_id = core_project_dict[core_project_number]
        s = f'addNIHGrantCoreProject(id: \\"{graph_id}\\", coreProject: [\\"{core_project_graph_id}\\"])'
        util.send_mutation(s, server)
        # print(core_project_graph_id, flush=True)


def fix_dates(d):
    dd = ''
    if d!='':
        toks = d.split('/')
        dd = toks[2] + '-' + pad(toks[0]) + '-' + pad(toks[1])
    return dd

def pad(s):
    if len(s)==1:
        s = '0' + s
    return s


def init_server(name):
    server = name
    util.initialize_graph(server)


def read_grants(server,core_project_dict):
    oranization_dict = {}
    ic_dict = {}
    pi_dict = {}

    grants = read_grants_file()
    count = 0
    for grant in grants:
        create_NIHGrant(grant, server, oranization_dict, ic_dict, pi_dict, core_project_dict)
        count += 1
        if count % 100==0:
            print(count,flush=True)

def read_publications_file():
    xlsx_file = Path('data', 'Pubhl_Export_thin.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)
    pubs = []
    for row in sheet.iter_rows(min_row=2):
        pub = {}
        for cell in row:
            pub[col_names[cell.col_idx-1]] =  util.replace_characters(cell.value)
        pubs.append(pub)
    return pubs

def read_publications(server,core_project_dict):
    pubs = read_publications_file()
    reference_dict: dict = {}
    journal_dict: dict = {}
    author_dict: dict= {}

    counter = 0
    for pub in pubs:
        # print(pub['Core Project Number'], pub['PMID'],counter)
        s, ref_id = write_references(pub['PMID'], reference_dict, journal_dict, author_dict)
        if ref_id != '':
            if s!='':
                util.send_mutation(s, server)
            core_project_number = pub['Core Project Number']
            if core_project_number in core_project_dict:
                graph_id = core_project_dict[core_project_number]
                s = f'addCoreProjectPublications(id: \\"{graph_id}\\", publications: [\\"{ref_id}\\"])'
                util.send_mutation(s, server)
            else:
                print('missing:',core_project_number)
        counter += 1

        if counter % 100==0:
            print(counter)

# ref_29320745
#  use "pip install -r requirements.txt" to get required modules

def read_ct_file():
    trials = []
    with open('data/CT_Export_04Feb2021_064551.csv', 'r') as read_obj:
        csv_dict_reader = csv.DictReader(read_obj)
        for row in csv_dict_reader:
            ct_obj = {'core_project_number':row["Core Project Number"],'nct_id':row["ClinicalTrials.gov Identifier"]}
            trials.append(ct_obj)
    return trials

# createClinicalTrial(
# brief_summary: String!
# brief_title: String!
# conditions: [String]!
# detailed_description: String!
# drugs: [String]!
# id: ID!
# nct_id: String!
# official_title: String!
# phases: [String]!
# status: String!
# status_date: String!
# study_type: String!): String
def create_clinical_trial(ct_obj):
    graph_id: str = util.get_unique_id('clinical_trial_')
    drugs_str = ''
    drugs_lcss = ''
    phase_lcss = ''
    condition_lcss = ''

    conditions_str = ''
    for condition in ct_obj['conditions']:
        conditions_str += f'\\"{condition}\\",'
        condition_lcss += ' ' + condition
    lc = ct_obj["nct_id"] + ' ' + ct_obj["brief_title"]

    for drug in ct_obj['drugs']:
        drugs_str += f'\\"{drug}\\",'
        drugs_lcss += ' ' + drug

    phase_str = ''
    for phase in ct_obj['phases']:
        phase_str += f'\\"{phase}\\",'
        phase_lcss += ' ' + phase

    drugs_lcss = drugs_lcss.lower()
    phase_lcss = phase_lcss.lower()
    condition_lcss = condition_lcss.lower()
    lc = lc.lower() + drugs_lcss + phase_lcss + condition_lcss

    s: str = f'{graph_id}: createClinicalTrial( brief_summary: \\"{ct_obj["brief_summary"]}\\",brief_title: \\"{ct_obj["brief_title"]}\\", conditions: [{conditions_str}], ' \
             f'detailed_description: \\"{ct_obj["detailed_description"]}\\", drugs: [{drugs_str}], id: \\"{graph_id}\\",' \
             f'nct_id:\\"{ct_obj["nct_id"]}\\",official_title:\\"{ct_obj["official_title"]}\\",phases: [{phase_str}],' \
             f'status:\\"{ct_obj["status"]}\\", status_date:\\"{ct_obj["status_date"]}\\", study_type:\\"{ct_obj["study_type"]}\\", ' \
             f'lower_case_search_string: \\"{lc}\\",phase_lower_case_search_string: \\"{phase_lcss}\\"condition_lower_case_search_string: \\"{condition_lcss}\\"drug_lower_case_search_string: \\"{drugs_lcss}\\"  ),'
    print(graph_id)
    util.send_mutation(s, server)
    return graph_id


def read_clinical_trials(server, core_project_dict):
    trials = read_ct_file()
    for trial in trials:
        ct_obj = get_trial_info(trial['nct_id'])
        core_project_number = trial['core_project_number']
        core_project_graph_id = core_project_dict[core_project_number]
        print(ct_obj)
        ct_graph_id = create_clinical_trial(ct_obj)
#         addClinicalTrialCoreProject(coreProject: [ID!]!id: ID!): String
        s = f'addClinicalTrialCoreProject(id: \\"{ct_graph_id}\\", coreProject: [\\"{core_project_graph_id}\\"])'
        util.send_mutation(s, server)

# createUser(
# email: String!
# firstInitial: String!
# id: ID!
# password: String!
# surname: String!
# user_name: String!): String
def add_user(server,email,initial,surname,user_name,password):
    graph_id = util.get_unique_id('user_')
    lc = initial.lower() + ' ' + surname.lower()
    s = f'createUser(email: \\"{email}\\",firstInitial: \\"{initial}\\", id: \\"{graph_id}\\", lower_case_search_string:\\"{lc}\\", password: \\"{password}\\", surname: \\"{surname}\\", user_name: \\"{user_name}\\")'
    util.send_mutation(s, server)


def add_users(server):
    add_user(server,'mglynias@gmail.com','M','Glynias','mglynias','pepper')
    add_user(server, 'cfranken@rush.edu', 'C', 'Frankenberger', 'cfrankenberger', 'casey')


def add_non_grant_ct():
    trials = retreive_trials()
    for nct in trials:
        ct_obj = get_trial_info(nct)
        create_clinical_trial(ct_obj)


if __name__ == '__main__':
    # server = 'localhost'
    server = '164.90.131.102'
    init_server(server)
    add_users(server)
    core_project_dict = {}
    read_grants(server,core_project_dict)
    read_clinical_trials(server,core_project_dict)
    add_non_grant_ct()
    read_publications(server,core_project_dict)



